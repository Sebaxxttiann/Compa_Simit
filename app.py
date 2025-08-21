from flask import Flask, render_template_string, request, jsonify, send_file
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
import os
import time
import json
import threading
from datetime import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import platform

# Configuración para Railway (Linux)
def configurar_chrome_para_railway():
    options = Options()
    
    if platform.system() == "Linux":
        options.add_argument('--headless')  # Sin interfaz gráfica
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-gpu')
        options.add_argument('--remote-debugging-port=9222')
        options.add_argument('--disable-extensions')
        options.add_argument('--disable-plugins')
        options.add_argument('--window-size=1920,1080')
        # Railway tiene Chrome preinstalado
        options.binary_location = "/usr/bin/google-chrome"
    else:
        # Para desarrollo local
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-gpu')
    
    return options

app = Flask(__name__)

# Variables globales para el progreso
progreso_actual = {
    'estado': 'idle',
    'mensaje': 'Listo para iniciar',
    'placa_actual': '',
    'total': 0,
    'procesadas': 0,
    'porcentaje': 0,
    'resultados': [],
    'archivo_excel': ''
}

class SimitScraper:
    def __init__(self):
        self.resultados = []
        self.driver = None
    
    def actualizar_progreso(self, mensaje, placa_actual='', total=0, procesadas=0):
        global progreso_actual
        
        porcentaje = 0
        if total > 0:
            porcentaje = round((procesadas / total) * 100, 1)
            porcentaje = min(porcentaje, 100)
        
        progreso_actual.update({
            'mensaje': mensaje,
            'placa_actual': placa_actual,
            'total': total,
            'procesadas': procesadas,
            'porcentaje': porcentaje
        })

    def esperar_carga_simple(self, driver):
        """Espera simplificada para carga de página"""
        try:
            WebDriverWait(driver, 20).until(
                lambda d: d.execute_script("return document.readyState") == "complete"
            )
            time.sleep(5)  # Espera fija para que cargue todo
            return True
        except:
            time.sleep(5)
            return True

    def detectar_multas_mejorada(self, driver, placa):
        """Detección CORREGIDA de multas usando los selectores reales de SIMIT"""
        try:
            time.sleep(2)
            
            # MÉTODO 1: Buscar tabla específica de SIMIT
            try:
                # Buscar por el ID exacto de la tabla de multas
                tabla_multas = driver.find_element(By.ID, "multaTable")
                tbody = tabla_multas.find_element(By.TAG_NAME, "tbody")
                filas = tbody.find_elements(By.TAG_NAME, "tr")
                
                # Filtrar filas que no son de "sin resultados"
                filas_con_multas = []
                for fila in filas:
                    texto_fila = fila.text.strip().lower()
                    # Si la fila tiene datos reales de multa (no mensajes de "sin resultados")
                    if texto_fila and not any(palabra in texto_fila for palabra in [
                        'no se encontraron', 'sin multas', 'no hay multas', 'no tiene multas'
                    ]):
                        # Verificar que tiene celdas con datos reales
                        celdas = fila.find_elements(By.TAG_NAME, "td")
                        if len(celdas) >= 6:  # Una multa real debe tener al menos 6 columnas
                            filas_con_multas.append(fila)
                
                if len(filas_con_multas) > 0:
                    print(f"✅ MULTAS DETECTADAS: {len(filas_con_multas)} multa(s) en tabla")
                    return True, len(filas_con_multas)
                else:
                    print("✅ TABLA ENCONTRADA pero SIN MULTAS")
                    return False, 0
                    
            except Exception as e:
                print(f"No se encontró tabla #multaTable: {e}")
            
            # MÉTODO 2: Buscar mensaje específico de "sin multas"
            try:
                # Buscar elementos que indiquen que no hay multas
                sin_multas_elementos = driver.find_elements(By.XPATH, 
                    "//*[contains(text(), 'No se encontraron') or contains(text(), 'sin multas') or contains(text(), 'No hay multas')]")
                
                if len(sin_multas_elementos) > 0:
                    print("✅ SIN MULTAS - Mensaje encontrado")
                    return False, 0
                    
            except:
                pass
            
            # MÉTODO 3: Analizar texto general de la página
            texto_pagina = driver.page_source.lower()
            
            # Mensajes específicos que indican NO hay multas
            sin_multas_frases = [
                "no se encontraron multas",
                "sin multas registradas", 
                "no hay multas",
                "no tiene multas",
                "sin infracciones",
                "no se encontraron infracciones"
            ]
            
            for frase in sin_multas_frases:
                if frase in texto_pagina:
                    print(f"✅ SIN MULTAS - Frase encontrada: '{frase}'")
                    return False, 0
            
            # MÉTODO 4: Buscar indicadores positivos de multas
            if any(palabra in texto_pagina for palabra in [
                "valor a pagar", "cobro coactivo", "secretaría", "infracción"
            ]):
                print("✅ POSIBLES MULTAS - Indicadores encontrados")
                return True, 1
            
            # Por defecto, asumir que no hay multas
            print("✅ SIN MULTAS - No se encontraron indicadores")
            return False, 0
            
        except Exception as e:
            print(f"❌ Error en detección: {e}")
            return False, 0

    def extraer_detalles_multas(self, driver, placa):
        """Extracción CORREGIDA de detalles usando la estructura real de SIMIT"""
        detalles = ""
        try:
            # Buscar la tabla específica de multas
            tabla_multas = driver.find_element(By.ID, "multaTable")
            tbody = tabla_multas.find_element(By.TAG_NAME, "tbody")
            filas = tbody.find_elements(By.TAG_NAME, "tr")
            
            multa_count = 0
            for fila in filas:
                try:
                    texto_fila = fila.text.strip()
                    if not texto_fila or any(palabra in texto_fila.lower() for palabra in [
                        'no se encontraron', 'sin multas'
                    ]):
                        continue
                    
                    celdas = fila.find_elements(By.TAG_NAME, "td")
                    if len(celdas) >= 6:
                        multa_count += 1
                        detalles += f"=== MULTA {multa_count} ===\n"
                        
                        # Extraer datos según la estructura real de SIMIT
                        if len(celdas) > 0:  # Tipo y número
                            tipo_cell = celdas[0].text.strip()
                            detalles += f"Tipo: {tipo_cell}\n"
                        
                        if len(celdas) > 1:  # Notificación
                            notif_cell = celdas[1].text.strip()
                            detalles += f"Notificación: {notif_cell}\n"
                        
                        if len(celdas) > 2:  # Placa
                            placa_cell = celdas[2].text.strip()
                            detalles += f"Placa: {placa_cell}\n"
                        
                        if len(celdas) > 3:  # Secretaría
                            secretaria_cell = celdas[3].text.strip()
                            detalles += f"Secretaría: {secretaria_cell}\n"
                        
                        if len(celdas) > 4:  # Infracción
                            infraccion_cell = celdas[4].text.strip()
                            detalles += f"Infracción: {infraccion_cell}\n"
                        
                        if len(celdas) > 5:  # Estado
                            estado_cell = celdas[5].text.strip()
                            detalles += f"Estado: {estado_cell}\n"
                        
                        if len(celdas) > 6:  # Valor
                            valor_cell = celdas[6].text.strip()
                            detalles += f"Valor: {valor_cell}\n"
                        
                        if len(celdas) > 7:  # Valor a pagar
                            valor_pagar_cell = celdas[7].text.strip()
                            detalles += f"Valor a pagar: {valor_pagar_cell}\n"
                        
                        detalles += "\n"
                        
                except Exception as e:
                    print(f"Error extrayendo fila de multa: {e}")
                    continue
                    
        except Exception as e:
            print(f"Error extrayendo detalles de multas: {e}")
            detalles = "No se pudieron extraer detalles específicos"
            
        return detalles.strip() if detalles.strip() else "Sin detalles disponibles"

    def tomar_captura_simple(self, placa, driver):
        """Captura simple de pantalla"""
        try:
            if not os.path.exists("capturas"):
                os.makedirs("capturas")
            
            screenshot_path = f"capturas/{placa}_{datetime.now().strftime('%H%M%S')}.png"
            
            driver.execute_script("window.scrollTo(0, 0);")
            time.sleep(1)
            
            driver.save_screenshot(screenshot_path)
            
            if os.path.exists(screenshot_path):
                return screenshot_path
            else:
                return "Sin captura"
                
        except:
            return "Sin captura"

    def buscar_placas(self, placas):
        global progreso_actual
        
        try:
            progreso_actual['estado'] = 'processing'
            self.actualizar_progreso("Iniciando proceso...", total=len(placas), procesadas=0)
            
            # Railway maneja Chrome automáticamente
            service = Service()  # Sin especificar ruta de chromedriver
            options = configurar_chrome_para_railway()
            
            self.driver = webdriver.Chrome(service=service, options=options)
            
            # Solo maximizar si no es headless
            if platform.system() != "Linux":
                self.driver.maximize_window()
            
            self.actualizar_progreso("Navegando a SIMIT...", total=len(placas), procesadas=0)
            self.driver.get("https://www.fcm.org.co/simit/#/home-public")
            
            self.esperar_carga_simple(self.driver)
            
            # Procesar cada placa
            for idx, placa in enumerate(placas):
                try:
                    self.actualizar_progreso(f"Procesando: {placa}", placa, len(placas), idx)
                    
                    # Cerrar popups
                    try:
                        popup = WebDriverWait(self.driver, 2).until(
                            EC.presence_of_element_located((By.CLASS_NAME, "swal2-popup"))
                        )
                        cerrar_btn = self.driver.find_element(By.CLASS_NAME, "swal2-confirm")
                        cerrar_btn.click()
                        time.sleep(1)
                    except:
                        pass

                    # Buscar placa
                    campo_placa = WebDriverWait(self.driver, 10).until(
                        EC.element_to_be_clickable((By.ID, "txtBusqueda"))
                    )
                    
                    campo_placa.clear()
                    time.sleep(0.5)
                    campo_placa.send_keys(placa)
                    time.sleep(1)
                    campo_placa.send_keys("\n")
                    
                    # Esperar resultados
                    time.sleep(8)
                    
                    # Detectar multas CORREGIDO
                    tiene_multas, num_multas = self.detectar_multas_mejorada(self.driver, placa)
                    
                    # Extraer detalles si hay multas
                    detalle_multas = ""
                    if tiene_multas:
                        self.actualizar_progreso(f"Extrayendo detalles de {placa}...", placa, len(placas), idx)
                        detalle_multas = self.extraer_detalles_multas(self.driver, placa)
                    
                    # Tomar captura
                    screenshot_path = self.tomar_captura_simple(placa, self.driver)
                    
                    estado_multas = "Sí" if tiene_multas else "No"
                    self.resultados.append((placa, estado_multas, "Éxito", screenshot_path, detalle_multas))
                    
                    # Actualizar progreso
                    procesadas_actual = idx + 1
                    self.actualizar_progreso(f"Completada: {placa} ({estado_multas} multas)", placa, len(placas), procesadas_actual)
                    
                    time.sleep(2)
                    
                except Exception as e:
                    procesadas_actual = idx + 1
                    self.actualizar_progreso(f"Error en {placa}", placa, len(placas), procesadas_actual)
                    self.resultados.append((placa, "Error", "Error", "Sin captura", str(e)))
            
            # Generar Excel
            self.actualizar_progreso("Generando Excel...", total=len(placas), procesadas=len(placas))
            archivo_excel = self.guardar_resultados_en_excel()
            
            if archivo_excel and os.path.exists(archivo_excel):
                progreso_actual.update({
                    'estado': 'completed',
                    'resultados': self.resultados,
                    'archivo_excel': archivo_excel,
                    'porcentaje': 100,
                    'procesadas': len(placas),
                    'total': len(placas),
                    'mensaje': 'Proceso completado. Excel listo para descarga.'
                })
            else:
                raise Exception("Error generando Excel")
            
        except Exception as e:
            progreso_actual.update({
                'estado': 'error',
                'mensaje': f"Error: {str(e)}",
                'porcentaje': 0
            })
        finally:
            try:
                if self.driver:
                    self.driver.quit()
            except:
                pass

    def guardar_resultados_en_excel(self):
        try:
            if not os.path.exists("reportes_excel"):
                os.makedirs("reportes_excel")
                
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            archivo = f"reportes_excel/reporte_simit_{timestamp}.xlsx"
            
            wb = Workbook()
            ws1 = wb.active
            ws1.title = "Control de Multas"

            # Colores y estilos - RESTAURADOS
            verde_oscuro = "1F7246"
            verde_claro = "C6E0B4"
            rojo_claro = "FFE6E6"
            
            # Configuración de columnas - MEJORADA
            ws1.column_dimensions['A'].width = 15
            ws1.column_dimensions['B'].width = 15
            ws1.column_dimensions['C'].width = 15
            ws1.column_dimensions['D'].width = 35
            ws1.column_dimensions['E'].width = 50  # Para detalles

            # Título principal - RESTAURADO
            ws1.merge_cells('A1:E2')
            titulo = ws1.cell(row=1, column=1, value="FORMATO DE CONTROL DE MULTAS DE TRÁNSITO")
            titulo.font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
            titulo.alignment = Alignment(horizontal="center", vertical="center")
            titulo.fill = PatternFill(start_color=verde_oscuro, end_color=verde_oscuro, fill_type="solid")

            # Fecha del reporte - RESTAURADA
            ws1.merge_cells('A3:E3')
            fecha = ws1.cell(row=3, column=1, value=f"Reporte generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
            fecha.font = Font(name='Arial', size=10, italic=True)
            fecha.alignment = Alignment(horizontal="right")

            # Encabezados - RESTAURADOS con columna de detalles
            encabezados = ["Placa", "Estado Multas", "Resultado", "Evidencia", "Detalles"]
            for col, encabezado in enumerate(encabezados, 1):
                celda = ws1.cell(row=4, column=col, value=encabezado)
                celda.font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
                celda.fill = PatternFill(start_color=verde_oscuro, end_color=verde_oscuro, fill_type="solid")
                celda.alignment = Alignment(horizontal="center", vertical="center")

            # Datos - MEJORADOS
            for idx, (placa, tiene_multa, resultado, captura, detalle_multas) in enumerate(self.resultados, 5):
                # Color de fila según estado - RESTAURADO
                if tiene_multa == "Sí":
                    fill_color = rojo_claro
                elif resultado == "Error":
                    fill_color = "FFCCCC"
                else:
                    fill_color = verde_claro if idx % 2 == 0 else "FFFFFF"
                
                # Datos de la fila - INCLUYENDO DETALLES
                datos_fila = [placa, tiene_multa, resultado, "Ver imagen adjunta", detalle_multas or "Sin detalles"]
                
                for col_idx, valor in enumerate(datos_fila, 1):
                    celda = ws1.cell(row=idx, column=col_idx, value=valor)
                    celda.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    
                    if col_idx == 1:  # Placa
                        celda.alignment = Alignment(horizontal="center")
                    elif col_idx == 2 and tiene_multa == "Sí":  # Estado Multas
                        celda.font = Font(color="FF0000", bold=True)
                        celda.alignment = Alignment(horizontal="center")
                    elif col_idx == 5:  # Detalles - RESTAURADO
                        celda.alignment = Alignment(wrap_text=True, vertical="top")
                
                # Agregar imagen - RESTAURADO
                if captura != "Sin captura" and os.path.exists(captura):
                    try:
                        img = Image(captura)
                        img.width = 300
                        img.height = 150
                        ws1.row_dimensions[idx].height = 120
                        ws1.add_image(img, f"D{idx}")
                    except Exception as e:
                        print(f"Error agregando imagen: {e}")

            # Guardar archivo
            wb.save(archivo)
            
            if os.path.exists(archivo) and os.path.getsize(archivo) > 1000:
                return archivo
            else:
                return None
            
        except Exception as e:
            print(f"Error generando Excel: {e}")
            return None

# RUTAS FLASK
@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)

@app.route('/iniciar_proceso', methods=['POST'])
def iniciar_proceso():
    global progreso_actual
    
    try:
        data = request.get_json()
        placas_texto = data.get('placas', '')
        placas = [placa.strip().upper() for placa in placas_texto.split('\n') if placa.strip()]
        
        if not placas:
            return jsonify({'error': 'No se ingresaron placas válidas'}), 400
        
        if progreso_actual['estado'] == 'processing':
            return jsonify({'error': 'Ya hay un proceso en ejecución'}), 400
        
        # Reiniciar progreso
        progreso_actual = {
            'estado': 'idle',
            'mensaje': 'Iniciando...',
            'placa_actual': '',
            'total': len(placas),
            'procesadas': 0,
            'porcentaje': 0,
            'resultados': [],
            'archivo_excel': ''
        }
        
        scraper = SimitScraper()
        thread = threading.Thread(target=scraper.buscar_placas, args=(placas,))
        thread.daemon = True
        thread.start()
        
        return jsonify({'success': True, 'mensaje': 'Proceso iniciado', 'total_placas': len(placas)})
        
    except Exception as e:
        return jsonify({'error': f'Error: {str(e)}'}), 500

@app.route('/progreso')
def obtener_progreso():
    global progreso_actual
    return jsonify(progreso_actual.copy())

@app.route('/descargar_excel')
def descargar_excel():
    global progreso_actual
    
    archivo_excel = progreso_actual.get('archivo_excel', '')
    
    if archivo_excel and os.path.exists(archivo_excel):
        try:
            return send_file(
                archivo_excel, 
                as_attachment=True,
                download_name=f"reporte_simit_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        except Exception as e:
            return jsonify({'error': f'Error enviando archivo: {str(e)}'}), 500
    else:
        return jsonify({'error': 'No hay archivo disponible'}), 404

# HTML TEMPLATE
HTML_TEMPLATE = '''
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>SIMIT Scraper</title>
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }
        
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            padding: 20px;
        }
        
        .container {
            max-width: 800px;
            margin: 0 auto;
            background: white;
            border-radius: 20px;
            box-shadow: 0 20px 40px rgba(0,0,0,0.1);
            overflow: hidden;
        }
        
        .header {
            background: linear-gradient(135deg, #1F7246 0%, #2E8B57 100%);
            color: white;
            padding: 30px;
            text-align: center;
        }
        
        .header h1 {
            font-size: 2.5em;
            margin-bottom: 10px;
        }
        
        .content {
            padding: 40px;
        }
        
        .input-group {
            margin-bottom: 30px;
        }
        
        .input-group label {
            display: block;
            font-weight: bold;
            margin-bottom: 10px;
            color: #333;
            font-size: 1.1em;
        }
        
        .placas-input {
            width: 100%;
            height: 200px;
            padding: 15px;
            border: 2px solid #ddd;
            border-radius: 10px;
            font-size: 16px;
            font-family: monospace;
            resize: vertical;
        }
        
        .placas-input:focus {
            outline: none;
            border-color: #1F7246;
        }
        
        .btn {
            background: linear-gradient(135deg, #1F7246 0%, #2E8B57 100%);
            color: white;
            border: none;
            padding: 15px 30px;
            font-size: 18px;
            border-radius: 10px;
            cursor: pointer;
            width: 100%;
            font-weight: bold;
        }
        
        .btn:hover:not(:disabled) {
            transform: translateY(-2px);
        }
        
        .btn:disabled {
            background: #ccc;
            cursor: not-allowed;
            transform: none;
        }
        
        .progress-container {
            display: none;
            margin-top: 30px;
            padding: 25px;
            background: #f8f9fa;
            border-radius: 15px;
            border-left: 5px solid #1F7246;
        }
        
        .progress-bar {
            width: 100%;
            height: 35px;
            background: #e9ecef;
            border-radius: 20px;
            overflow: hidden;
            margin: 20px 0;
            position: relative;
        }
        
        .progress-fill {
            height: 100%;
            background: linear-gradient(90deg, #1F7246, #2E8B57, #20c997);
            width: 0%;
            transition: width 0.5s ease-out;
            display: flex;
            align-items: center;
            justify-content: center;
            color: white;
            font-weight: bold;
            font-size: 16px;
        }
        
        .progress-info {
            display: grid;
            grid-template-columns: 1fr 1fr 1fr;
            gap: 15px;
            margin-top: 20px;
        }
        
        .progress-item {
            background: white;
            padding: 15px;
            border-radius: 10px;
            text-align: center;
        }
        
        .progress-item strong {
            display: block;
            color: #1F7246;
            font-size: 1.3em;
            margin-bottom: 5px;
        }
        
        .results-container {
            display: none;
            margin-top: 30px;
            padding: 25px;
            background: linear-gradient(135deg, #e8f5e8 0%, #d4edda 100%);
            border-radius: 15px;
            border: 2px solid #1F7246;
            text-align: center;
        }
        
        .download-btn {
            background: linear-gradient(135deg, #28a745 0%, #20c997 100%);
            margin-top: 20px;
            padding: 15px 40px;
        }
        
        .status-message {
            padding: 15px;
            margin: 15px 0;
            border-radius: 10px;
            font-weight: bold;
        }
        
        .status-success {
            background: #d4edda;
            color: #155724;
            border: 2px solid #b8dabc;
        }
        
        .status-error {
            background: #f8d7da;
            color: #721c24;
            border: 2px solid #f1aeb5;
        }
        
        .status-info {
            background: #d1ecf1;
            color: #0c5460;
            border: 2px solid #abdde5;
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🚗 SIMIT Scraper</h1>
            <p>Sistema de Control de Multas</p>
        </div>
        
        <div class="content">
            <div class="input-group">
                <label for="placas">Ingrese las placas (una por línea):</label>
                <textarea 
                    id="placas" 
                    class="placas-input" 
                    placeholder="ABC123&#10;DEF456&#10;GHI789"
                >ABC123
DEF456</textarea>
            </div>
            
            <button id="iniciarBtn" class="btn" onclick="iniciarProceso()">
                🚀 Iniciar Búsqueda
            </button>
            
            <div id="progressContainer" class="progress-container">
                <h3>Progreso del Proceso</h3>
                <div class="progress-bar">
                    <div id="progressFill" class="progress-fill">0%</div>
                </div>
                <div id="statusMessage" class="status-message status-info">
                    <span id="statusText">Iniciando proceso...</span>
                </div>
                <div class="progress-info">
                    <div class="progress-item">
                        <strong id="placaActual">-</strong>
                        <span>Placa Actual</span>
                    </div>
                    <div class="progress-item">
                        <strong id="contador">0 / 0</strong>
                        <span>Procesadas</span>
                    </div>
                    <div class="progress-item">
                        <strong id="estadoGeneral">Iniciando</strong>
                        <span>Estado</span>
                    </div>
                </div>
            </div>
            
            <div id="resultsContainer" class="results-container">
                <h3>🎉 ¡Proceso Completado!</h3>
                <p>El reporte Excel ha sido generado con todos los detalles de multas.</p>
                <button id="downloadBtn" class="btn download-btn" onclick="descargarExcel()">
                    📥 Descargar Reporte Excel
                </button>
            </div>
        </div>
    </div>

    <script>
        let intervalId = null;
        let procesoIniciado = false;
        
        function iniciarProceso() {
            const placasTexto = document.getElementById('placas').value.trim();
            
            if (!placasTexto) {
                alert('Por favor, ingrese al menos una placa.');
                return;
            }
            
            const placasArray = placasTexto.split('\\n').filter(p => p.trim());
            
            if (placasArray.length === 0) {
                alert('No se encontraron placas válidas.');
                return;
            }
            
            if (!confirm(`¿Iniciar búsqueda para ${placasArray.length} placa(s)?`)) {
                return;
            }
            
            procesoIniciado = true;
            
            // Cambiar UI
            const btn = document.getElementById('iniciarBtn');
            btn.disabled = true;
            btn.textContent = '⏳ Procesando...';
            
            document.getElementById('progressContainer').style.display = 'block';
            document.getElementById('resultsContainer').style.display = 'none';
            
            // Resetear progreso
            const progressFill = document.getElementById('progressFill');
            progressFill.style.width = '0%';
            progressFill.textContent = '0%';
            
            // Iniciar proceso
            fetch('/iniciar_proceso', {
                method: 'POST',
                headers: {
                    'Content-Type': 'application/json',
                },
                body: JSON.stringify({
                    placas: placasTexto
                })
            })
            .then(response => response.json())
            .then(data => {
                if (data.success) {
                    intervalId = setInterval(actualizarProgreso, 1000);
                } else {
                    throw new Error(data.error || 'Error desconocido');
                }
            })
            .catch(error => {
                alert('Error al iniciar proceso: ' + error.message);
                resetearUI();
            });
        }
        
        function actualizarProgreso() {
            if (!procesoIniciado) return;
            
            fetch('/progreso')
            .then(response => response.json())
            .then(data => {
                const porcentaje = Math.max(0, Math.min(100, data.porcentaje || 0));
                
                // Actualizar barra
                const progressFill = document.getElementById('progressFill');
                progressFill.style.width = porcentaje + '%';
                progressFill.textContent = porcentaje.toFixed(1) + '%';
                
                // Actualizar información
                document.getElementById('placaActual').textContent = data.placa_actual || '-';
                document.getElementById('contador').textContent = `${data.procesadas || 0} / ${data.total || 0}`;
                document.getElementById('estadoGeneral').textContent = data.estado || 'Procesando';
                
                // Actualizar mensaje
                const statusMessage = document.getElementById('statusMessage');
                const statusText = document.getElementById('statusText');
                statusText.textContent = data.mensaje || 'Procesando...';
                
                // Cambiar estilo según estado
                statusMessage.className = 'status-message';
                if (data.estado === 'completed') {
                    statusMessage.classList.add('status-success');
                    progressFill.style.width = '100%';
                    progressFill.textContent = '100%';
                    
                    clearInterval(intervalId);
                    procesoIniciado = false;
                    resetearUI();
                    
                    setTimeout(() => {
                        document.getElementById('resultsContainer').style.display = 'block';
                    }, 1000);
                    
                } else if (data.estado === 'error') {
                    statusMessage.classList.add('status-error');
                    clearInterval(intervalId);
                    procesoIniciado = false;
                    resetearUI();
                } else {
                    statusMessage.classList.add('status-info');
                }
            })
            .catch(error => {
                console.error('Error polling:', error);
            });
        }
        
        function resetearUI() {
            const btn = document.getElementById('iniciarBtn');
            btn.disabled = false;
            btn.textContent = '🚀 Iniciar Búsqueda';
        }
        
        function descargarExcel() {
            const btn = document.getElementById('downloadBtn');
            btn.disabled = true;
            btn.textContent = '📥 Descargando...';
            
            fetch('/descargar_excel')
            .then(response => {
                if (!response.ok) {
                    throw new Error('Error en la descarga');
                }
                return response.blob();
            })
            .then(blob => {
                const url = window.URL.createObjectURL(blob);
                const a = document.createElement('a');
                a.href = url;
                a.download = `reporte_simit_${new Date().toISOString().slice(0,10)}.xlsx`;
                document.body.appendChild(a);
                a.click();
                document.body.removeChild(a);
                window.URL.revokeObjectURL(url);
                
                btn.disabled = false;
                btn.textContent = '📥 Descargar Reporte Excel';
                
                alert('¡Archivo descargado exitosamente!');
            })
            .catch(error => {
                btn.disabled = false;
                btn.textContent = '📥 Descargar Reporte Excel';
                alert('Error al descargar: ' + error.message);
            });
        }
    </script>
</body>
</html>
'''

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))
    app.run(debug=False, host='0.0.0.0', port=port)

